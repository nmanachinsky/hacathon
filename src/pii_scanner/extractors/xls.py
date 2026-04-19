"""XLS / XLSX → текст."""

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path

from ..types import TextChunk


def _extract_xlsx(path: Path) -> Iterable[TextChunk]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    for sheet in wb.worksheets:
        rows: list[str] = []
        header_added = False
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None and str(c) != ""]
            if not cells:
                continue
            if not header_added:
                rows.append(" ".join(cells))
                header_added = True
            else:
                rows.append("; ".join(cells))
        if rows:
            yield TextChunk(text="\n".join(rows), locator=f"sheet={sheet.title}")
    wb.close()


def _extract_xls(path: Path) -> Iterable[TextChunk]:
    import xlrd

    book = xlrd.open_workbook(str(path), formatting_info=False, on_demand=True)
    for sheet_name in book.sheet_names():
        sheet = book.sheet_by_name(sheet_name)
        rows: list[str] = []
        for r in range(sheet.nrows):
            cells = [str(sheet.cell_value(r, c)) for c in range(sheet.ncols)
                     if sheet.cell_value(r, c) not in (None, "")]
            if cells:
                rows.append("; ".join(cells))
        if rows:
            yield TextChunk(text="\n".join(rows), locator=f"sheet={sheet_name}")
        book.unload_sheet(sheet_name)


def extract(path: Path) -> Iterable[TextChunk]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        yield from _extract_xlsx(path)
    else:
        yield from _extract_xls(path)
